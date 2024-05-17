'''
functions file
'''

from openpyxl import load_workbook
def father_of(file):
    workbook = load_workbook(filename=file)     #reading exel file
    workbook.sheetnames
    sheet = workbook.active

    name = [(cell.value).strip() for cell in sheet['B']][1:]    #read columns B and D then make list two lists from them
    level = [cell.value for cell in sheet['D']][1:]

    result_dict = dict()    # an empty dictionary with fathers as keys and children as values
    for i_index,i in enumerate(level):
        if i == 0:      #find main product with level 0 and then find its children
            temp_list = []      #a temporary list to append children in them
            for j_index,j in enumerate(level):
                if j==1:
                    temp_list.append(name[j_index])
            result_dict[name[i_index]] = temp_list      # put children list into their father value
        else:
            temp_list = []
            for j_index,j in enumerate(level[i_index+1:]):
                if j == i+1:
                    temp_list.append(name[j_index+i_index+1])
                elif i==j:      #use to avoide unexpected data
                    break
            result_dict[name[i_index]] = temp_list

    return result_dict

def son_of(file):
    workbook = load_workbook(filename=file)
    workbook.sheetnames
    sheet = workbook.active

    name = [(cell.value).strip() for cell in sheet['B']][::-1][:-1]     #name and level lists but in reverse format
    level = [cell.value for cell in sheet['D']][::-1][:-1]

    result_dict = dict()
    for i_index,i in enumerate(level):
        for j_index,j in enumerate(level[i_index+1:]):
            if i-1 == j:
                if name[i_index] in result_dict:        #to find children with two father
                    result_dict[name[i_index]].append(name[j_index+i_index+1])
                else:
                    result_dict[name[i_index]] = [name[j_index+i_index+1]]
                break
    return result_dict

def total_numb(file,number):
    workbook = load_workbook(filename=file)
    workbook.sheetnames
    sheet = workbook.active

    father_dict = father_of(file)       #use the father to children dictionary in func 1
    needed_parts = {list(father_dict.keys())[0]:number}     #an empty dictionary but the first item is the final product
    # names as keys and the numbers as values
    name_number = dict(zip([cell.value.strip() for cell in sheet['B']][1:],[cell.value for cell in sheet['E']][1:]))
    # names as keys and the units as values
    unit_dict = dict(zip([cell.value.strip() for cell in sheet['B']][1:],[cell.value for cell in sheet['F']][1:]))

    for key,value in name_number.items():
        if value is None:       #find none cells in exel file
            name_number[key] = 0
    for i in father_dict:
        for j in father_dict[i]:
            needed_parts[j] = needed_parts[i] * name_number[j]

    for parts in needed_parts:      #add unit to numbers in needed_parts dictionary
        needed_parts[parts] = f'{needed_parts[parts]} {unit_dict[parts]}'
    return needed_parts
